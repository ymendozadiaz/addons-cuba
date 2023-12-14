import tkinter as tk
import customtkinter as ctk
import sqlite3
import qrcode
from tkinter import *
from tkinter import messagebox, Menu, filedialog
import openpyxl
from PIL import ImageTk, Image
import os
import psutil
import wmi
from openpyxl import Workbook
import pandas as pd
import getpass
import platform
import datetime
import socket

root = ctk.CTk()
root.geometry('960x600')
root.title('Progen QR')
root.resizable=False

mode = ctk.set_appearance_mode("dark")

def color_choice():
    switch_state = activar.get()
    if switch_state:
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("green")
    else:
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")


frame_mode = ctk.CTkFrame(root)
frame_mode.pack(padx=10, pady=10)

activar = ctk.CTkSwitch(frame_mode, command=color_choice, text='Modo Claro')
activar.pack(padx=20, pady=20, side='left')


conn = sqlite3.connect('almacedb.db')
c = conn.cursor()

c.execute('''CREATE TABLE IF NOT EXISTS user
                    (id INTEGER PRIMARY KEY, user TEXT, password TEXT, role TEXT)''')

def login_window():  
    conn = sqlite3.connect('almacedb.db')
    c = conn.cursor()
                    
        # Crear tabla si no existe
    c.execute('''CREATE TABLE IF NOT EXISTS user
                    (id INTEGER PRIMARY KEY, user TEXT, password TEXT, role TEXT)''')  

            
    # Función para verificar el inicio de sesión

def verificar_login():
    username = entry_username.get()
    password = entry_password.get()  

    # Buscar el usuario en la tabla "user"
    c.execute("SELECT * FROM user WHERE user=?", (username,))        
    result = c.fetchone()

    if username == "admin" and password == "12345678":
            root.destroy()
            main_window_admin()
    else:
        if result:
        # Verificar la contraseña
            if result[1] == username:
                if result[2] == password:
                    root.destroy()
                    main_window_user()
                else:
                    messagebox.showerror('Error', 'Contraseña no válida.')
                    entry_username.delete(0, END)
                    entry_password.delete(0, END)
            else:
                messagebox.showerror('Error', 'Credenciales incorrectas.')
                entry_username.delete(0, END)
                entry_password.delete(0, END)
        else:
            messagebox.showerror('Error', 'No existen permisos para las credenciales proporcionadas.')
            entry_username.delete(0, END)
            entry_password.delete(0, END)

def main_window_admin():
    root = ctk.CTk()
    root.geometry('960x600')
    root.title('Progen QR')
    root.resizable=False
    def Inicio_page():
        conn = sqlite3.connect('almacedb.db')
        c = conn.cursor()

            # Crear tabla si no existe
        c.execute('''CREATE TABLE IF NOT EXISTS productos
                    (id INTEGER PRIMARY KEY, nombre TEXT, departamento TEXT, motherboard TEXT, cpu TEXT,  ram TEXT, hdd TEXT)''')

        def importar_desde_excel():
            archivo = filedialog.askopenfile(filetypes=(("Archivos Excel", "*.xlsx"),))
            if archivo:
                wb = openpyxl.load_workbook(archivo.name)
                hoja = wb.active

                conn = sqlite3.connect('almacedb.db')
                cursor = conn.cursor()
                
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    nombre = fila[0]
                    departamento = fila[1]
                    motherboard = fila[2]
                    cpu = fila[3]
                    ram = fila[4]
                    hdd = fila[5]
                    
                    cursor.execute("INSERT INTO productos (nombre, departamento, motherboard, cpu, ram, hdd) VALUES (?, ?, ?, ?, ?, ?)",
                        (nombre, departamento, motherboard, cpu, ram, hdd))
                
                conn.commit()
                conn.close()

                messagebox.showinfo("Importado", "Los datos han sido importados exitosamente.")
            else:
                messagebox.showinfo("Importado", "No se realizó la importación de los datos")
            update_listbox()

            cursor.close()

        def export_to_excel():
            # Obtener los datos ingresados por el usuario
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos")
            num_records = cursor.fetchone()[0]

            if num_records > 0:
                df = pd.read_sql_query("SELECT * FROM productos", conn)
                filename = filedialog.asksaveasfilename(initialdir="/", title="Guardar como", defaultextension='.xlsx', filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
                df.to_excel(filename, index=False)
                messagebox.showinfo("Información", "El reporte en excel se generó con éxito")
            else:
                messagebox.showwarning("Advertencia", "No hay registros para exportar")

            cursor.close()

        def clean_entry():
            nombre_entry.delete(0, END)
            departamento_entry.delete(0, END)
            motherboard_entry.delete(0, END)
            cpu_entry.delete(0, END)
            ram_entry.delete(0, END)        
            hdd_entry.delete(0, END)

        # Función para insertar un nuevo producto
        def insert_producto():
            nombre = nombre_entry.get()
            departamento = departamento_entry.get()
            motherboard = motherboard_entry.get()
            cpu = cpu_entry.get()
            ram = ram_entry.get()
            hdd = hdd_entry.get()
            c.execute("INSERT INTO productos (nombre, departamento, motherboard, cpu, ram, hdd) VALUES (?, ?, ?, ?, ?, ?)",
                        (nombre, departamento, motherboard, cpu, ram, hdd))
            conn.commit()
            update_listbox()
            messagebox.showinfo("Información", "El registro fue agregado con éxito")
            clean_entry()

        # Función para actualizar un producto
        def update_producto():
            selected_producto = listbox.curselection()[0]
            nombre = nombre_entry.get()
            departamento = departamento_entry.get()
            motherboard = motherboard_entry.get()
            cpu = cpu_entry.get()
            ram = ram_entry.get()
            hdd = hdd_entry.get()
            producto_id = listbox.get(selected_producto)[0]
            c.execute("UPDATE productos SET nombre=?, departamento=?, motherboard=?, cpu=?, ram=?, hdd=? WHERE id=?",
                        (nombre, departamento, motherboard, cpu, ram, hdd, producto_id))
            conn.commit()
            update_listbox()
            messagebox.showinfo("Información", "El registro fue actualizado con éxito")

        # Función para eliminar un producto
        def delete_producto():
            selected_producto = listbox.curselection()[0]
            producto_id = listbox.get(selected_producto)[0]
            c.execute("DELETE FROM productos WHERE id=?", (producto_id,))
            conn.commit()
            update_listbox()
            messagebox.showinfo("Información", "El registro fue eliminado de la base de datos")

        # Función para mostrar los productos en la lista
        def update_listbox():
            listbox.delete(0, END)
            for row in c.execute("SELECT * FROM productos"):
                listbox.insert(END, row)
                count_registros()

        # Función para generar código QR
        def generate_qr():
            selected_producto = listbox.curselection()[0]
            producto_id = listbox.get(selected_producto)[0]
            nombre = listbox.get(selected_producto)[1]
            departamento = listbox.get(selected_producto)[2]
            motherboard = listbox.get(selected_producto)[3]
            cpu = listbox.get(selected_producto)[4]
            ram = listbox.get(selected_producto)[5]
            hdd = listbox.get(selected_producto)[6]
            qr_data = f"Nombre: {nombre}\nDepartamento: {departamento}\nMotherboard: {motherboard}\nCpu: {cpu}\nRam: {ram}\nhdd: {hdd}"
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(qr_data)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            filename = filedialog.asksaveasfilename(initialdir="/", title="Guardar como", defaultextension='.png', filetypes=(("Archivos png", "*.png") , ("Archivos jpg", "*.jpg"), ("Todos los archivos", "*.*")))
            img.save(filename)
            messagebox.showinfo("Información", message="Código QR generado y guardado correctamente.")

        def count_registros():
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos")
            num_registros = cursor.fetchone()[0]
            cantidad_label.configure(text=f" Total PC Registradas: {num_registros} ")

        def buscar_producto():
            nombre = nombre_buscar_entry.get()
            c.execute("SELECT * FROM productos WHERE nombre=?", (nombre,))        
            rows = c.fetchall()
            if len(rows) > 0:
                listbox.delete(0, END)
                for row in rows:
                    listbox.insert(END, row)
            else:
                messagebox.showinfo("Información", "No se encontro la PC con ese nombre")

        def count_registros():
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos")
            num_registros = cursor.fetchone()[0]
            cantidad_label.configure(text=f" Total PC Registradas: {num_registros} ")

        def buscar_producto():
            nombre = nombre_buscar_entry.get()
            c.execute("SELECT * FROM productos WHERE nombre=?", (nombre,))        
            rows = c.fetchall()
            if len(rows) > 0:
                listbox.delete(0, END)
                for row in rows:
                    listbox.insert(END, row)
            else:
                messagebox.showinfo("Información", "No se encontro la PC con ese nombre")

        nombre_label = ctk.CTkLabel(main_frame, text="Generador de QR", font=("Stencil", 50), width=25, height=5, text_color="Grey")
        nombre_label.pack(fill=BOTH)  

        search_frame = ctk.CTkFrame(main_frame) 
        cantidad_label = ctk.CTkLabel(search_frame, text=" Total PC Registradas: 0 ", font=("bold", 14))
        cantidad_label.configure(bg_color='green', text_color='white', corner_radius=50)
        cantidad_label.pack()

        nombre_buscar_entry = ctk.CTkEntry(search_frame, placeholder_text="PC a Buscar: ")
        nombre_buscar_entry.pack(padx=10, pady=10)

        buscar_button = ctk.CTkButton(search_frame, text="Buscar", border_color="grey", border_width=2, command=buscar_producto)
        buscar_button.pack(padx=10, pady=10)
        search_frame.pack(side="top", fill=BOTH, pady=5, padx=5)

        entry_frame = ctk.CTkFrame(main_frame)
        nombre_entry = ctk.CTkEntry(entry_frame, placeholder_text="Nombre de la PC")
        nombre_entry.pack(pady=10, padx=10)

        departamento_entry = ctk.CTkEntry(entry_frame, placeholder_text="Departamento")
        departamento_entry.pack(pady=10, padx=10)

        motherboard_entry = ctk.CTkEntry(entry_frame, placeholder_text="Modelo de motherboard")
        motherboard_entry.pack(pady=10, padx=10)

        cpu_entry = ctk.CTkEntry(entry_frame, placeholder_text="Modelo de CPU")
        cpu_entry.pack(pady=10, padx=10)

        ram_entry = ctk.CTkEntry(entry_frame, placeholder_text="Total GB de RAM ")
        ram_entry.pack(pady=10, padx=10)

        hdd_entry = ctk.CTkEntry(entry_frame, placeholder_text="Total GB del HDD ")
        hdd_entry.pack(pady=10, padx=10)
        entry_frame.pack(side=LEFT, pady=10, padx=10)

        btn_frame = ctk.CTkFrame(main_frame)
        add_button = ctk.CTkButton(btn_frame, text="Agregar", command=insert_producto, border_color="grey", border_width=2)
        add_button.pack(pady=10, padx=10)

        update_button = ctk.CTkButton(btn_frame, text="Actualizar", command=update_producto, border_color="grey", border_width=2)
        update_button.pack(pady=10, padx=10)

        delete_button = ctk.CTkButton(btn_frame, text="Eliminar", command=delete_producto, border_color="grey", border_width=2)
        delete_button.pack(pady=10, padx=10)

        qr_button = ctk.CTkButton(btn_frame, text="Generar QR", command=generate_qr, border_color="grey", border_width=2)
        qr_button.pack(pady=10, padx=10)

        excel_button = ctk.CTkButton(btn_frame, text="Exportar", command=export_to_excel, border_color="grey", border_width=2)
        excel_button.pack(pady=10, padx=10)
        btn_frame.pack(side=LEFT, pady=10, padx=10)

        Listbox_frame = ctk.CTkFrame(main_frame) 
            # Crear lista de productos
        listbox = Listbox(Listbox_frame, height=20, width=200, border=2)
        listbox.pack(padx=8, pady=8, side=LEFT)
        ScrollVert = Scrollbar(Listbox_frame, command=listbox.yview)
        ScrollVert.pack(side=LEFT)
        listbox.config(yscrollcommand=ScrollVert.set)
        update_listbox()
        Listbox_frame.pack(fill=BOTH, pady=10)

    def Propiedades_page():
        Propiedades_frame = ctk.CTkFrame(main_frame)

        lb = ctk.CTkLabel(Propiedades_frame, text='Propiedades', font=("Stencil", 50), width=25, height=5, text_color="Grey")
        lb.pack(padx=10, pady=10, fill=BOTH)

        scrollable_frame = ctk.CTkScrollableFrame(Propiedades_frame, )
        scrollable_frame.pack(fill="both", expand=True)

        def get_system_info():
                c = wmi.WMI()
                pc_name = socket.gethostname()
                system_info = platform.uname()
                motherboard = c.Win32_BaseBoard()[0]
                cpu = c.Win32_Processor()[0]
                ram = c.Win32_PhysicalMemory()[0]
                memory = psutil.virtual_memory()
                drives = psutil.disk_partitions()[0]
                hdd = c.Win32_DiskDrive()[0]
                fecha_hora_actual = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                info_dict = {
                    'Pc name': pc_name,
                    'Motherboard': motherboard.Manufacturer,
                    'Modelo': motherboard.Product,
                    'Serial': motherboard.SerialNumber,
                    'Version BIOS': motherboard.Version,
                    'Fabricante del BIOS': motherboard.Name,
                    'Procesador': cpu.Name,
                    'Serial CPU': cpu.ProcessorId.strip(),
                    'Socket': cpu.SocketDesignation,
                    'Frecuencia': f"{cpu.MaxClockSpeed}MHz",
                    'Total RAM': memory.total,
                    'RAM Disponible': memory.available,
                    'RAM Usada': memory.used,
                    'Fabricante RAM': ram.Manufacturer,
                    'Serial RAM': ram.SerialNumber,
                    'Fabricante HDD': hdd.Manufacturer,
                    'Serial HDD': hdd.SerialNumber,
                    'HDD': drives.fstype,
                    'HDD Usage': psutil.disk_usage(drives.device).used,
                    'HDD Total': psutil.disk_usage(drives.device).total,
                    'Arquitectura del sistema': system_info.machine,
                    'Usuario': getpass.getuser(),
                    'Sistema operativo': f"{platform.system()} {platform.release()}",
                    'Fecha y hora actual': fecha_hora_actual
                }

                return info_dict

        def add_todo():
            info_dict = get_system_info()
            inf_lbl = ""
            for key, value in info_dict.items():
                inf_lbl += f"{key}: \n{value}\n\n"

            label = ctk.CTkLabel(scrollable_frame, text=inf_lbl)
            label.pack()

            conn = sqlite3.connect('almacedb.db')
            c = conn.cursor()

            c.execute('''CREATE TABLE IF NOT EXISTS propiedades
                        (id INTEGER PRIMARY KEY AUTOINCREMENT, Nombre_PC TEXT, Motherboard TEXT, Modelo TEXT, Serial_1 TEXT, Version_BIOS TEXT, Fabricante_BIOS TEXT, CPU TEXT, Serial_CPU TEXT, Frecuencia TEXT, Socket TEXT, Fabricante_RAM TEXT, Serial_RAM TEXT, Total_RAM TEXT, RAM_Disponible TEXT, RAM_Usada TEXT, Fabricante_HDD TEXT, Serial_HDD TEXT, Tipo_de_archivo_HDD TEXT, HDD_Libre TEXT, HDD_Total TEXT, Arquitectura_sistema TEXT, Usuario TEXT, Sistema_operativo TEXT, Fecha_y_hora_actual TEXT)''')

            # Verificar si la PC ya existe en la tabla
            c.execute("SELECT COUNT(*) FROM propiedades WHERE Nombre_PC=?", (info_dict['Pc name'],))
            pc_count = c.fetchone()[0]

            if pc_count == 0:
                c.execute("INSERT INTO propiedades (Nombre_PC, Motherboard, Modelo, Serial_1, Version_BIOS, Fabricante_BIOS, CPU, Serial_CPU, Frecuencia, Socket, Fabricante_RAM, Serial_RAM, Total_RAM, RAM_Disponible, RAM_Usada, Fabricante_HDD, Serial_HDD, Tipo_de_archivo_HDD, HDD_Libre, HDD_Total, Arquitectura_sistema, Usuario, Sistema_operativo, Fecha_y_hora_actual) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (info_dict['Pc name'], info_dict['Motherboard'], info_dict['Modelo'], info_dict['Serial'], info_dict['Version BIOS'], info_dict['Fabricante del BIOS'],
                        info_dict['Procesador'], info_dict['Serial CPU'], info_dict['Frecuencia'], info_dict['Socket'], info_dict['Fabricante RAM'], info_dict['Serial RAM'],
                        info_dict['Total RAM'], info_dict['RAM Disponible'], info_dict['RAM Usada'], info_dict['Fabricante HDD'], info_dict['Serial HDD'], info_dict['HDD'],
                        info_dict['HDD Usage'], info_dict['HDD Total'], info_dict['Arquitectura del sistema'], info_dict['Usuario'], info_dict['Sistema operativo'],
                        info_dict['Fecha y hora actual']))

                conn.commit()


            df = pd.DataFrame(info_dict, index=[0])
            filename = filedialog.asksaveasfilename(initialdir="/", title="Guardar como", defaultextension='.xlsx',
                                                    filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
            df.to_excel(filename, index=False)

        def count_registros():
            conn = sqlite3.connect('almacedb.db')
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM propiedades")
            num_registros = cursor.fetchone()[0]
            lb.configure(text=f" Total PC Registradas: {num_registros} ")

        def export_to_excel():
                conn = sqlite3.connect('almacedb.db')
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM propiedades")
                num_records = cursor.fetchone()[0]

                if num_records > 0:
                    df = pd.read_sql_query("SELECT * FROM propiedades", conn)
                    filename = filedialog.asksaveasfilename(initialdir="/", title="Guardar como", defaultextension='.xlsx', filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
                    df.to_excel(filename, index=False)
                    messagebox.showinfo("Información", "Reporte en excel se generado con éxito")
                else:
                    messagebox.showwarning("Advertencia", "No hay registros para exportar")
                    

                cursor.close()

        def buscar_producto():
            # Limpiar el contenido del scrollable_frame
            for widget in scrollable_frame.winfo_children():
                widget.destroy()

            count_registros()
            conn = sqlite3.connect('almacedb.db')
            c = conn.cursor()

            search_term = nombre_buscar_entry.get()

            c.execute("SELECT * FROM propiedades WHERE Nombre_PC LIKE ?", ('%' + search_term + '%',))
            results = c.fetchall()

            conn.close()

            if results:
                for result in results:
                    table_text = ""
                    table_text += "Nombre_PC: " + result[1] + "\n"
                    table_text += "Motherboard: " + result[2] + "\n"
                    table_text += "Modelo: " + result[3] + "\n"
                    table_text += "Serial_1: " + result[4] + "\n"
                    table_text += "Version_BIOS: " + result[5] + "\n"
                    table_text += "Fabricante_BIOS: " + result[6] + "\n"
                    table_text += "CPU: " + result[7] + "\n"
                    table_text += "Serial_CPU: " + result[8] + "\n"
                    table_text += "Frecuencia: " + result[9] + "\n"
                    table_text += "Socket: " + result[10] + "\n"
                    table_text += "Fabricante_RAM: " + result[11] + "\n"
                    table_text += "Serial_RAM: " + result[12] + "\n"
                    table_text += "Total_RAM: " + result[13] + "\n"
                    table_text += "RAM_Disponible: " + result[14] + "\n"
                    table_text += "RAM_Usada: " + result[15] + "\n"
                    table_text += "Fabricante_HDD: " + result[16] + "\n"
                    table_text += "Serial_HDD: " + result[17] + "\n"
                    table_text += "Tipo_de_archivo_HDD: " + result[18] + "\n"
                    table_text += "HDD_Libre: " + result[19] + "\n"
                    table_text += "HDD_Total: " + result[20] + "\n"
                    table_text += "Arquitectura_sistema: " + result[21] + "\n"
                    table_text += "Usuario: " + result[22] + "\n"
                    table_text += "Sistema_operativo: " + result[23] + "\n"
                    table_text += "Fecha_y_hora_actual: " + result[24] + "\n"
                    table_text += "\n"

                    table = ctk.CTkLabel(scrollable_frame, text=table_text)
                    table.pack()
            else:
                label = ctk.CTkLabel(scrollable_frame, text="No se encontraron resultados.")
                label.pack()

        frame_search = ctk.CTkFrame(Propiedades_frame)
        frame_search.pack(side='left', padx=20, pady=20)

        frame_panel = ctk.CTkFrame(Propiedades_frame)
        frame_panel.pack(side='left', padx=20, pady=20)

        nombre_buscar_entry = ctk.CTkEntry(frame_search, placeholder_text="Buscar por Nombre", width=200, border_color="grey", border_width=2,)
        nombre_buscar_entry.pack(side="left", padx=10, pady=10)

        btn_buscar = ctk.CTkButton(frame_search, text="Buscar", width=50, command=buscar_producto, border_color="grey", border_width=2,)
        btn_buscar.pack(side="right", padx=10, pady=10)

        report_button = ctk.CTkButton(frame_panel, text="Generar Reporte", border_color="grey", border_width=2, width=140, height=28, command=add_todo)
        report_button.pack(side="right", padx=10, pady=10)

        export_button = ctk.CTkButton(frame_panel, text="Exportar BD", border_color="grey", border_width=2, width=140, height=28, command=export_to_excel)
        export_button.pack(side="right", padx=10, pady=10)

        Propiedades_frame.pack(fill="both", expand=True, pady=10)

    def Registro_page():
        lbl_registro = ctk.CTkLabel(main_frame, text="Registro de nuevo usuario", font=("Stencil", 50), width=25, height=5, text_color="Grey")
        lbl_registro.pack(fill="both", side=TOP, padx=10, pady=50)

        user_entry = ctk.CTkEntry(main_frame, placeholder_text="Escribe el nombre del nuevo usuario", height=50, width=250, corner_radius=50, border_color="grey", border_width=2)
        user_entry.pack(padx=10, pady=10)

        pass_entry = ctk.CTkEntry(main_frame, placeholder_text="Contraseña para el nuevo usuario", show='*', height=50, width=250, corner_radius=50, border_color="grey", border_width=2)
        pass_entry.pack(padx=10, pady=10) 

        btn_frame = ctk.CTkFrame(main_frame)
        btn_frame.pack(pady=20, padx=20)

        def main():
            new_user = user_entry.get()
            new_pass = pass_entry.get()
            new_role = 'user'

            conn = sqlite3.connect('almacedb.db')
            c = conn.cursor()

            c.execute("INSERT INTO user (user, password, role) VALUES (?, ?, ?)",
                    (new_user, new_pass, new_role))
            conn.commit()
            messagebox.showinfo("Información", "El registro del nuevo usuario fue guardado con éxito")

            conn.close()

        def search_and_change_password():
            user_to_search = user_entry.get()

            conn = sqlite3.connect('almacedb.db')
            c = conn.cursor()
            
            c.execute("SELECT user FROM user WHERE user = ?", (user_to_search,))
            result = c.fetchone()

            if result:
                new_pass = pass_entry.get()
                c.execute("UPDATE user SET password = ? WHERE user = ?", (new_pass, user_to_search))
                conn.commit()
                messagebox.showinfo("Información", "Contraseña cambiada exitosamente")
            else:
                messagebox.showinfo("Información", "Usuario no encontrado")

            conn.close()

        search_button = ctk.CTkButton(btn_frame, text="Buscar y Cambiar Contraseña", command=search_and_change_password, border_color="grey", border_width=2)
        search_button.pack(side=LEFT, padx=2, pady=2)

        add_button = ctk.CTkButton(btn_frame, text="Guardar", command=main, border_color="grey", border_width=2)
        add_button.pack(side=LEFT, padx=5, pady=5)

    def Acerca_page():
        Acerca_frame = ctk.CTkFrame(main_frame)

        lbl_title = ctk.CTkLabel(Acerca_frame, text="Acerca del Programa", font=("Stencil", 50), width=25, height=5, text_color="Grey")
        lbl_title.pack(fill="both", padx=10, pady=10)


        textinf = """Este programa fue creado con el objetivo de facilitar la manera de obtener \n
        información sobre las computadoras por medio de un QR. \n
        Programa compilado en noviembre del 2023, versión 3.0. \n
        Programado por: Yosbel Mendoza Diaz. \n
        CI: 86051216107\n
        Dirección: Calle 10 Edif. 7 Apto. 14, Morón. \n"""

        lbl_info = ctk.CTkLabel(Acerca_frame, text=textinf, font=('Canabria', 12), width=10, height=3, anchor='n')
        lbl_info.pack(padx=10, pady=10)

        Acerca_frame.pack(pady=20, side='top')

    def hide_indicators():
        Inicio_indicate.config(bg='black')
        Propiedades_indicate.config(bg='black')
        Registro_indicate.config(bg='black')
        Acerca_indicate.config(bg='black')

    def delete_page():
        for frame in main_frame.winfo_children():
            frame.destroy()

    def indicate(lb, page):
        hide_indicators()
        lb.config(bg='green')
        delete_page()
        page()

    options_frame = ctk.CTkFrame(root, bg_color='black', fg_color='black')

    Inicio_btn = ctk.CTkButton(options_frame, text='Generar QR', font=('Bold', 20), hover_color='green', fg_color='black', width=150, height=40, bg_color='Black', command=lambda:indicate(Inicio_indicate, Inicio_page))
    Inicio_btn.place(x=10, y=50)
    Inicio_indicate = tk.Label(options_frame, text='', bg='black')
    Inicio_indicate.place(x=3, y=50, width=5, height=40)

    Propiedades = ctk.CTkButton(options_frame, text='Propiedades', font=('Bold', 20), hover_color='green', fg_color='black', width=150, height=40, bg_color='Black', command=lambda:indicate(Propiedades_indicate, Propiedades_page))
    Propiedades.place(x=10, y=100)
    Propiedades_indicate = tk.Label(options_frame, text='', bg='black')
    Propiedades_indicate.place(x=3, y=100, width=5, height=40)

    Registro_btn = ctk.CTkButton(options_frame, text='Registro', font=('Bold', 20), hover_color='green', fg_color='black', width=150, height=40, bg_color='Black', command=lambda:indicate(Registro_indicate, Registro_page))
    Registro_btn.place(x=10, y=150)
    Registro_indicate = tk.Label(options_frame, text='', bg='black')
    Registro_indicate.place(x=3, y=150, width=5, height=40)

    Acerca_btn = ctk.CTkButton(options_frame, text='Acerca', font=('Bold', 20), hover_color='green', fg_color='black', width=150, height=40, bg_color='Black', command=lambda:indicate(Acerca_indicate, Acerca_page))
    Acerca_btn.place(x=10, y=200)
    Acerca_indicate = tk.Label(options_frame, text='', bg='black')
    Acerca_indicate.place(x=3, y=200, width=5, height=40)

    options_frame.pack(side=tk.LEFT)
    options_frame.pack_propagate(False)
    options_frame.configure(width=150, height=600)

    main_frame = ctk.CTkFrame(root)
    main_frame.pack(side='left')
    main_frame.pack_propagate(False)
    main_frame.configure(width=810, height=600)
    root.mainloop()

def main_window_user():
    root = ctk.CTk()
    root.geometry('960x600')
    root.title('Progen QR')
    root.resizable=False
    def Inicio_page():
        conn = sqlite3.connect('almacedb.db')
        c = conn.cursor()

            # Crear tabla si no existe
        c.execute('''CREATE TABLE IF NOT EXISTS productos
                    (id INTEGER PRIMARY KEY, nombre TEXT, departamento TEXT, motherboard TEXT, cpu TEXT,  ram TEXT, hdd TEXT)''')

        def importar_desde_excel():
            archivo = filedialog.askopenfile(filetypes=(("Archivos Excel", "*.xlsx"),))
            if archivo:
                wb = openpyxl.load_workbook(archivo.name)
                hoja = wb.active

                conn = sqlite3.connect('almacedb.db')
                cursor = conn.cursor()
                
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    nombre = fila[0]
                    departamento = fila[1]
                    motherboard = fila[2]
                    cpu = fila[3]
                    ram = fila[4]
                    hdd = fila[5]
                    
                    cursor.execute("INSERT INTO productos (nombre, departamento, motherboard, cpu, ram, hdd) VALUES (?, ?, ?, ?, ?, ?)",
                        (nombre, departamento, motherboard, cpu, ram, hdd))
                
                conn.commit()
                conn.close()

                messagebox.showinfo("Importado", "Los datos han sido importados exitosamente.")
            else:
                messagebox.showinfo("Importado", "No se realizó la importación de los datos")
            update_listbox()

            cursor.close()

        def export_to_excel():
            # Obtener los datos ingresados por el usuario
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos")
            num_records = cursor.fetchone()[0]

            if num_records > 0:
                df = pd.read_sql_query("SELECT * FROM productos", conn)
                filename = filedialog.asksaveasfilename(initialdir="/", title="Guardar como", defaultextension='.xlsx', filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
                df.to_excel(filename, index=False)
                messagebox.showinfo("Información", "El reporte en excel se generó con éxito")
            else:
                messagebox.showwarning("Advertencia", "No hay registros para exportar")

            cursor.close()

        def clean_entry():
            nombre_entry.delete(0, END)
            departamento_entry.delete(0, END)
            motherboard_entry.delete(0, END)
            cpu_entry.delete(0, END)
            ram_entry.delete(0, END)        
            hdd_entry.delete(0, END)

        # Función para insertar un nuevo producto
        def insert_producto():
            nombre = nombre_entry.get()
            departamento = departamento_entry.get()
            motherboard = motherboard_entry.get()
            cpu = cpu_entry.get()
            ram = ram_entry.get()
            hdd = hdd_entry.get()
            c.execute("INSERT INTO productos (nombre, departamento, motherboard, cpu, ram, hdd) VALUES (?, ?, ?, ?, ?, ?)",
                        (nombre, departamento, motherboard, cpu, ram, hdd))
            conn.commit()
            update_listbox()
            messagebox.showinfo("Información", "El registro fue agregado con éxito")
            clean_entry()

        # Función para actualizar un producto
        def update_producto():
            selected_producto = listbox.curselection()[0]
            nombre = nombre_entry.get()
            departamento = departamento_entry.get()
            motherboard = motherboard_entry.get()
            cpu = cpu_entry.get()
            ram = ram_entry.get()
            hdd = hdd_entry.get()
            producto_id = listbox.get(selected_producto)[0]
            c.execute("UPDATE productos SET nombre=?, departamento=?, motherboard=?, cpu=?, ram=?, hdd=? WHERE id=?",
                        (nombre, departamento, motherboard, cpu, ram, hdd, producto_id))
            conn.commit()
            update_listbox()
            messagebox.showinfo("Información", "El registro fue actualizado con éxito")

        # Función para eliminar un producto
        def delete_producto():
            selected_producto = listbox.curselection()[0]
            producto_id = listbox.get(selected_producto)[0]
            c.execute("DELETE FROM productos WHERE id=?", (producto_id,))
            conn.commit()
            update_listbox()
            messagebox.showinfo("Información", "El registro fue eliminado de la base de datos")

        # Función para mostrar los productos en la lista
        def update_listbox():
            listbox.delete(0, END)
            for row in c.execute("SELECT * FROM productos"):
                listbox.insert(END, row)
                count_registros()

        # Función para generar código QR
        def generate_qr():
            selected_producto = listbox.curselection()[0]
            producto_id = listbox.get(selected_producto)[0]
            nombre = listbox.get(selected_producto)[1]
            departamento = listbox.get(selected_producto)[2]
            motherboard = listbox.get(selected_producto)[3]
            cpu = listbox.get(selected_producto)[4]
            ram = listbox.get(selected_producto)[5]
            hdd = listbox.get(selected_producto)[6]
            qr_data = f"Nombre: {nombre}\nDepartamento: {departamento}\nMotherboard: {motherboard}\nCpu: {cpu}\nRam: {ram}\nhdd: {hdd}"
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(qr_data)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            filename = filedialog.asksaveasfilename(initialdir="/", title="Guardar como", defaultextension='.png', filetypes=(("Archivos png", "*.png") , ("Archivos jpg", "*.jpg"), ("Todos los archivos", "*.*")))
            img.save(filename)
            messagebox.showinfo("Información", message="Código QR generado y guardado correctamente.")

        def count_registros():
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM productos")
            num_registros = cursor.fetchone()[0]
            cantidad_label.configure(text=f" Total PC Registradas: {num_registros} ")

        def buscar_producto():
            nombre = nombre_buscar_entry.get()
            c.execute("SELECT * FROM productos WHERE nombre=?", (nombre,))        
            rows = c.fetchall()
            if len(rows) > 0:
                listbox.delete(0, END)
                for row in rows:
                    listbox.insert(END, row)
            else:
                messagebox.showinfo("Información", "No se encontro la PC con ese nombre")


        nombre_label = ctk.CTkLabel(main_frame, text="Generador de QR", font=("Stencil", 50), width=25, height=5, text_color="Grey")
        nombre_label.pack(fill=BOTH)  

        search_frame = ctk.CTkFrame(main_frame) 
        cantidad_label = ctk.CTkLabel(search_frame, text=" Total PC Registradas: 0 ", font=("bold", 14))
        cantidad_label.configure(bg_color='green', text_color='white', corner_radius=50)
        cantidad_label.pack()

        nombre_buscar_entry = ctk.CTkEntry(search_frame, placeholder_text="PC a Buscar: ")
        nombre_buscar_entry.pack(padx=10, pady=10)

        buscar_button = ctk.CTkButton(search_frame, text="Buscar", border_color="grey", border_width=2, command=buscar_producto)
        buscar_button.pack(padx=10, pady=10)
        search_frame.pack(side="top", fill=BOTH, pady=5, padx=5)

        entry_frame = ctk.CTkFrame(main_frame)
        nombre_entry = ctk.CTkEntry(entry_frame, placeholder_text="Nombre de la PC")
        nombre_entry.pack(pady=10, padx=10)

        departamento_entry = ctk.CTkEntry(entry_frame, placeholder_text="Departamento")
        departamento_entry.pack(pady=10, padx=10)

        motherboard_entry = ctk.CTkEntry(entry_frame, placeholder_text="Modelo de motherboard")
        motherboard_entry.pack(pady=10, padx=10)

        cpu_entry = ctk.CTkEntry(entry_frame, placeholder_text="Modelo de CPU")
        cpu_entry.pack(pady=10, padx=10)

        ram_entry = ctk.CTkEntry(entry_frame, placeholder_text="Total GB de RAM ")
        ram_entry.pack(pady=10, padx=10)

        hdd_entry = ctk.CTkEntry(entry_frame, placeholder_text="Total GB del HDD ")
        hdd_entry.pack(pady=10, padx=10)
        entry_frame.pack(side=LEFT, pady=10, padx=10)

        btn_frame = ctk.CTkFrame(main_frame)
        add_button = ctk.CTkButton(btn_frame, text="Agregar", command=insert_producto, border_color="grey", border_width=2)
        add_button.pack(pady=10, padx=10)

        update_button = ctk.CTkButton(btn_frame, text="Actualizar", command=update_producto, border_color="grey", border_width=2)
        update_button.pack(pady=10, padx=10)

        delete_button = ctk.CTkButton(btn_frame, text="Eliminar", command=delete_producto, border_color="grey", border_width=2)
        delete_button.pack(pady=10, padx=10)

        qr_button = ctk.CTkButton(btn_frame, text="Generar QR", command=generate_qr, border_color="grey", border_width=2)
        qr_button.pack(pady=10, padx=10)

        excel_button = ctk.CTkButton(btn_frame, text="Exportar", command=export_to_excel, border_color="grey", border_width=2)
        excel_button.pack(pady=10, padx=10)
        btn_frame.pack(side=LEFT, pady=10, padx=10)

        # Crear lista de productos
        Listbox_frame = ctk.CTkFrame(main_frame) 
        listbox = Listbox(Listbox_frame, height=20, width=200, border=2)
        listbox.pack(padx=10, pady=10, side=LEFT)
        ScrollVert = Scrollbar(Listbox_frame, command=listbox.yview)
        ScrollVert.pack(side=LEFT)
        listbox.config(yscrollcommand=ScrollVert.set)
        update_listbox()
        Listbox_frame.pack(fill=BOTH, pady=10)

    def Propiedades_page():
        Propiedades_frame = ctk.CTkFrame(main_frame)

        lb = ctk.CTkLabel(Propiedades_frame, text='Propiedades', font=("Stencil", 50), width=25, height=5, text_color="Grey")
        lb.pack(pady=10, fill=BOTH)

        scrollable_frame = ctk.CTkScrollableFrame(Propiedades_frame)
        scrollable_frame.pack(fill="both", expand=True)

        def get_system_info():
                c = wmi.WMI()
                pc_name = socket.gethostname()
                system_info = platform.uname()
                motherboard = c.Win32_BaseBoard()[0]
                cpu = c.Win32_Processor()[0]
                ram = c.Win32_PhysicalMemory()[0]
                memory = psutil.virtual_memory()
                drives = psutil.disk_partitions()[0]
                hdd = c.Win32_DiskDrive()[0]
                fecha_hora_actual = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                info_dict = {
                    'Pc name': pc_name,
                    'Motherboard': motherboard.Manufacturer,
                    'Modelo': motherboard.Product,
                    'Serial': motherboard.SerialNumber,
                    'Version BIOS': motherboard.Version,
                    'Fabricante del BIOS': motherboard.Name,
                    'Procesador': cpu.Name,
                    'Serial CPU': cpu.ProcessorId.strip(),
                    'Socket': cpu.SocketDesignation,
                    'Frecuencia': f"{cpu.MaxClockSpeed}MHz",
                    'Total RAM': memory.total,
                    'RAM Disponible': memory.available,
                    'RAM Usada': memory.used,
                    'Fabricante RAM': ram.Manufacturer,
                    'Serial RAM': ram.SerialNumber,
                    'Fabricante HDD': hdd.Manufacturer,
                    'Serial HDD': hdd.SerialNumber,
                    'HDD': drives.fstype,
                    'HDD Usage': psutil.disk_usage(drives.device).used,
                    'HDD Total': psutil.disk_usage(drives.device).total,
                    'Arquitectura del sistema': system_info.machine,
                    'Usuario': getpass.getuser(),
                    'Sistema operativo': f"{platform.system()} {platform.release()}",
                    'Fecha y hora actual': fecha_hora_actual
                }

                return info_dict

        def add_todo():
            info_dict = get_system_info()
            inf_lbl = ""
            for key, value in info_dict.items():
                inf_lbl += f"{key}: \n{value}\n\n"

            label = ctk.CTkLabel(scrollable_frame, text=inf_lbl)
            label.pack()

            conn = sqlite3.connect('almacedb.db')
            c = conn.cursor()

            c.execute('''CREATE TABLE IF NOT EXISTS propiedades
                        (id INTEGER PRIMARY KEY AUTOINCREMENT, Nombre_PC TEXT, Motherboard TEXT, Modelo TEXT, Serial_1 TEXT, Version_BIOS TEXT, Fabricante_BIOS TEXT, CPU TEXT, Serial_CPU TEXT, Frecuencia TEXT, Socket TEXT, Fabricante_RAM TEXT, Serial_RAM TEXT, Total_RAM TEXT, RAM_Disponible TEXT, RAM_Usada TEXT, Fabricante_HDD TEXT, Serial_HDD TEXT, Tipo_de_archivo_HDD TEXT, HDD_Libre TEXT, HDD_Total TEXT, Arquitectura_sistema TEXT, Usuario TEXT, Sistema_operativo TEXT, Fecha_y_hora_actual TEXT)''')

            # Verificar si la PC ya existe en la tabla
            c.execute("SELECT COUNT(*) FROM propiedades WHERE Nombre_PC=?", (info_dict['Pc name'],))
            pc_count = c.fetchone()[0]

            if pc_count == 0:
                c.execute("INSERT INTO propiedades (Nombre_PC, Motherboard, Modelo, Serial_1, Version_BIOS, Fabricante_BIOS, CPU, Serial_CPU, Frecuencia, Socket, Fabricante_RAM, Serial_RAM, Total_RAM, RAM_Disponible, RAM_Usada, Fabricante_HDD, Serial_HDD, Tipo_de_archivo_HDD, HDD_Libre, HDD_Total, Arquitectura_sistema, Usuario, Sistema_operativo, Fecha_y_hora_actual) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (info_dict['Pc name'], info_dict['Motherboard'], info_dict['Modelo'], info_dict['Serial'], info_dict['Version BIOS'], info_dict['Fabricante del BIOS'],
                        info_dict['Procesador'], info_dict['Serial CPU'], info_dict['Frecuencia'], info_dict['Socket'], info_dict['Fabricante RAM'], info_dict['Serial RAM'],
                        info_dict['Total RAM'], info_dict['RAM Disponible'], info_dict['RAM Usada'], info_dict['Fabricante HDD'], info_dict['Serial HDD'], info_dict['HDD'],
                        info_dict['HDD Usage'], info_dict['HDD Total'], info_dict['Arquitectura del sistema'], info_dict['Usuario'], info_dict['Sistema operativo'],
                        info_dict['Fecha y hora actual']))

                conn.commit()


            df = pd.DataFrame(info_dict, index=[0])
            filename = filedialog.asksaveasfilename(initialdir="/", title="Guardar como", defaultextension='.xlsx',
                                                    filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
            df.to_excel(filename, index=False)

        def count_registros():
            conn = sqlite3.connect('almacedb.db')
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM propiedades")
            num_registros = cursor.fetchone()[0]
            lb.configure(text=f" Total PC Registradas: {num_registros} ")

        def export_to_excel():
                conn = sqlite3.connect('almacedb.db')
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM propiedades")
                num_records = cursor.fetchone()[0]

                if num_records > 0:
                    df = pd.read_sql_query("SELECT * FROM propiedades", conn)
                    filename = filedialog.asksaveasfilename(initialdir="/", title="Guardar como", defaultextension='.xlsx', filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
                    df.to_excel(filename, index=False)
                    messagebox.showinfo("Información", "Reporte en excel se generado con éxito")
                else:
                    messagebox.showwarning("Advertencia", "No hay registros para exportar")
                    

                cursor.close()

        def buscar_producto():
            # Limpiar el contenido del scrollable_frame
            for widget in scrollable_frame.winfo_children():
                widget.destroy()

            count_registros()
            conn = sqlite3.connect('almacedb.db')
            c = conn.cursor()

            search_term = nombre_buscar_entry.get()

            c.execute("SELECT * FROM propiedades WHERE Nombre_PC LIKE ?", ('%' + search_term + '%',))
            results = c.fetchall()

            conn.close()

            if results:
                for result in results:
                    table_text = ""
                    table_text += "Nombre_PC: " + result[1] + "\n"
                    table_text += "Motherboard: " + result[2] + "\n"
                    table_text += "Modelo: " + result[3] + "\n"
                    table_text += "Serial_1: " + result[4] + "\n"
                    table_text += "Version_BIOS: " + result[5] + "\n"
                    table_text += "Fabricante_BIOS: " + result[6] + "\n"
                    table_text += "CPU: " + result[7] + "\n"
                    table_text += "Serial_CPU: " + result[8] + "\n"
                    table_text += "Frecuencia: " + result[9] + "\n"
                    table_text += "Socket: " + result[10] + "\n"
                    table_text += "Fabricante_RAM: " + result[11] + "\n"
                    table_text += "Serial_RAM: " + result[12] + "\n"
                    table_text += "Total_RAM: " + result[13] + "\n"
                    table_text += "RAM_Disponible: " + result[14] + "\n"
                    table_text += "RAM_Usada: " + result[15] + "\n"
                    table_text += "Fabricante_HDD: " + result[16] + "\n"
                    table_text += "Serial_HDD: " + result[17] + "\n"
                    table_text += "Tipo_de_archivo_HDD: " + result[18] + "\n"
                    table_text += "HDD_Libre: " + result[19] + "\n"
                    table_text += "HDD_Total: " + result[20] + "\n"
                    table_text += "Arquitectura_sistema: " + result[21] + "\n"
                    table_text += "Usuario: " + result[22] + "\n"
                    table_text += "Sistema_operativo: " + result[23] + "\n"
                    table_text += "Fecha_y_hora_actual: " + result[24] + "\n"
                    table_text += "\n"

                    table = ctk.CTkLabel(scrollable_frame, text=table_text)
                    table.pack()
            else:
                label = ctk.CTkLabel(scrollable_frame, text="No se encontraron resultados.")
                label.pack()

        frame_search = ctk.CTkFrame(Propiedades_frame)
        frame_search.pack(side='left', padx=20, pady=20)

        frame_panel = ctk.CTkFrame(Propiedades_frame)
        frame_panel.pack(side='left', padx=20, pady=20)

        nombre_buscar_entry = ctk.CTkEntry(frame_search, placeholder_text="Buscar por Nombre", width=200, border_color="grey", border_width=2)
        nombre_buscar_entry.pack(side="left", padx=10, pady=10)

        btn_buscar = ctk.CTkButton(frame_search, text="Buscar", width=50, command=buscar_producto, border_color="grey", border_width=2)
        btn_buscar.pack(side="right", padx=10, pady=10)

        report_button = ctk.CTkButton(frame_panel, text="Generar Reporte", border_color="grey", border_width=2, width=140, height=28, command=add_todo)
        report_button.pack(side="right", padx=10, pady=10)

        export_button = ctk.CTkButton(frame_panel, text="Exportar BD", border_color="grey", border_width=2, width=140, height=28, command=export_to_excel)
        export_button.pack(side="right", padx=10, pady=10)

        Propiedades_frame.pack(fill="both", expand=True, pady=10)

    def Acerca_page():
        Acerca_frame = ctk.CTkFrame(main_frame)

        lbl_title = ctk.CTkLabel(Acerca_frame, text="Acerca del Programa", font=("Stencil", 50), width=25, height=5, text_color="Grey")
        lbl_title.pack(fill="both", padx=10, pady=10)


        textinf = """Este programa fue creado con el objetivo de facilitar la manera de obtener \n
        información sobre las computadoras por medio de un QR. \n
        Programa compilado en noviembre del 2023, versión 3.0. \n
        Programado por: Yosbel Mendoza Diaz. \n
        CI: 86051216107\n
        Dirección: Calle 10 Edif. 7 Apto. 14, Morón. \n"""

        lbl_info = ctk.CTkLabel(Acerca_frame, text=textinf, font=('Canabria', 12), width=10, height=3, anchor='n')
        lbl_info.pack(padx=10, pady=10)

        Acerca_frame.pack(pady=20, side='top')

    def hide_indicators():
        Inicio_indicate.config(bg='black')
        Propiedades_indicate.config(bg='black')
        Acerca_indicate.config(bg='black')
    
    def delete_page():
        for frame in main_frame.winfo_children():
            frame.destroy()

    def indicate(lb, page):
        hide_indicators()
        lb.config(bg='green')
        delete_page()
        page()

    options_frame = ctk.CTkFrame(root, bg_color='black', fg_color='black')

    Inicio_btn = ctk.CTkButton(options_frame, text='Generar QR', font=('Bold', 20), hover_color='green', fg_color='black', width=150, height=40, bg_color='Black', command=lambda:indicate(Inicio_indicate, Inicio_page))
    Inicio_btn.place(x=10, y=50)
    Inicio_indicate = tk.Label(options_frame, text='', bg='black')
    Inicio_indicate.place(x=3, y=50, width=5, height=40)

    Propiedades = ctk.CTkButton(options_frame, text='Propiedades', font=('Bold', 20), hover_color='green', fg_color='black', width=150, height=40, bg_color='Black', command=lambda:indicate(Propiedades_indicate, Propiedades_page))
    Propiedades.place(x=10, y=100)
    Propiedades_indicate = tk.Label(options_frame, text='', bg='black')
    Propiedades_indicate.place(x=3, y=100, width=5, height=40)

    Acerca_btn = ctk.CTkButton(options_frame, text='Acerca', font=('Bold', 20), hover_color='green', fg_color='black', width=150, height=40, bg_color='Black', command=lambda:indicate(Acerca_indicate, Acerca_page))
    Acerca_btn.place(x=10, y=150)
    Acerca_indicate = tk.Label(options_frame, text='', bg='black')
    Acerca_indicate.place(x=3, y=150, width=5, height=40)

    options_frame.pack(side=tk.LEFT)
    options_frame.pack_propagate(False)
    options_frame.configure(width=150, height=600)

    main_frame = ctk.CTkFrame(root)
    main_frame.pack(side='left')
    main_frame.pack_propagate(False)
    main_frame.configure(width=810, height=600)
    root.mainloop()

image = ctk.CTkImage(light_image = Image.open("logo1.png"), size=(200, 200))
image_label = ctk.CTkLabel(root, image=image, text='')
image_label.pack(padx=10, pady=10)

label_username = ctk.CTkLabel(root, text='Usuario:')
label_username.pack()
entry_username = ctk.CTkEntry(root, placeholder_text="Escribe tu usuario", height=50, width=250, corner_radius=50)
entry_username.pack()

label_password = ctk.CTkLabel(root, text='Contraseña:')
label_password.pack()
entry_password = ctk.CTkEntry(root, placeholder_text="Escribe tu contraseña", height=50, width=250, show='*', corner_radius=50)
entry_password.pack()

button_login = ctk.CTkButton(root, text='Iniciar Sesión', corner_radius=10, command=verificar_login)
button_login.pack(pady=12, padx=10)

button_salir = ctk.CTkButton(root, text='Salir', corner_radius=10, command=root.quit, )
button_salir.pack(pady=10, padx=10)

root.mainloop()
